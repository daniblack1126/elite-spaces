# Elite Spaces — Source Audit XLSX Generator
# Run with: python generate_source_audit.py
# Requires: pip install openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

wb = Workbook()
ws = wb.active
ws.title = "Source Audit"

HDR_BG   = "1A1A2E"
HDR_FG   = "FFFFFF"
GREEN    = "D6F4E8"
YELLOW   = "FFF9C4"
RED_CELL = "FFE0E0"
ACCENT   = "4472C4"

thin = Side(style="thin", color="CCCCCC")
bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = [
    "Source Name",
    "RSS Feed (Yes/No + URL)",
    "Scrapable (Yes/No)",
    "Best Events Page URL",
    "Sample Events Found",
    "Notes / Obstacles",
]
col_widths = [22, 52, 18, 52, 60, 72]

for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font      = Font(name="Arial", bold=True, color=HDR_FG, size=10)
    cell.fill      = PatternFill("solid", fgColor=HDR_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = bdr
    ws.column_dimensions[get_column_letter(col_idx)].width = w

ws.row_dimensions[1].height = 30

rows = [
    (
        "Avenue Magazine",
        "Yes — avenuemagazine.com/category/events/feed/",
        "Yes",
        "avenuemagazine.com/category/events/",
        "Gala season roundups; society party coverage (retrospective posts)",
        "WordPress site; RSS feed confirmed accessible; events page covers past social events & party recaps, not a forward-facing calendar; HTTP to HTTPS redirects require handling in scraper",
        GREEN,
    ),
    (
        "Average Socialite",
        "Yes — averagesocialite.com/nyc-events?format=rss",
        "Yes",
        "averagesocialite.com/nyc",
        "Ray-Ban House (May 5-17); Park Ave Day (May 16); Made in NYC Week (May 1-7); Vogue Cafe (May 2); Asian Comedy Fest (May 5-7)",
        "Squarespace; structured URL scheme /nyc-events/YYYY/M/D/event-name; rich future-facing calendar; RSS confirmed accessible; excellent candidate for full automation",
        GREEN,
    ),
    (
        "CEW (Cosmetic Executive Women)",
        "No — none found",
        "Partial",
        "cew.org/events-overview/",
        "CEW Achiever Awards Luncheon (Apr 30, 2026); CEW Beauty Awards 2026; member meetups",
        "WordPress but events page redirects inconsistently; some events require CEW membership to register/view details; no RSS feed found; B2B beauty industry focus — not consumer-facing",
        YELLOW,
    ),
    (
        "Fairchild Live",
        "No — none found",
        "Yes",
        "events.fairchildlive.com/",
        "WWD Beauty CEO Summit 2026; LA Beauty Forum 2026; WWD Beauty Summit",
        "Industry conference/summit site — all events are professional B2B, not consumer-facing; page loads cleanly (no login wall); no RSS; JS-rendered event cards require browser scraping or structured HTML parsing",
        YELLOW,
    ),
    (
        "WWD",
        "Yes (news only) — wwd.com/feed/",
        "Partial",
        "wwd.com/feed/",
        "N/A — WWD is news coverage, not an event calendar",
        "General RSS feed works but contains articles only; no events-specific page (/category/events/ returns 404); would need keyword filtering to surface event mentions; partial paywall limits full content access",
        YELLOW,
    ),
    (
        "Stylus NYC",
        "No",
        "No",
        "N/A",
        "None found",
        "WRONG SOURCE: stylus.nyc is a music venue/record store on the Lower East Side, not a social or beauty calendar. Next.js, fully client-side rendered. Recommend removing from source list or clarifying intended source.",
        RED_CELL,
    ),
    (
        "The Select 7",
        "No",
        "No",
        "N/A",
        "None accessible",
        "Site is fully password-protected (HTTP 401 — Private Squarespace site); no public content visible at any URL. Requires direct relationship with publisher to access. Cannot be automated.",
        RED_CELL,
    ),
    (
        "Parasol Projects",
        "No — none found",
        "Yes",
        "parasolprojects.com/pop-ups",
        "Shopbop x Nike (Sep 2025); Barbour Sample Sale (Feb 2026); Sabina Savage (Feb 2026); Geel (Feb 2026)",
        "Clean, scrapable HTML with data-content-property attributes (easy to parse); pop-up brand activations only — not general social events; most currently listed events are past; worth monitoring for upcoming activations",
        YELLOW,
    ),
    (
        "NYC for Free",
        "Likely yes — nycforfree.co/events?format=rss",
        "Partial",
        "nycforfree.co/events",
        "Citi Concert Series TODAY (May 8+); BLACKPINK Deadline Pop-Up; Lincoln Center Big Umbrella Festival (Apr 10-26); Free Comic Book Day (May 2); Earth Day Festival (Apr 19)",
        "NOTE: Domain is .co not .com (brief lists .com — incorrect). Squarespace site; JS-heavy but Squarespace RSS endpoint likely works; strong free-event coverage; broad audience fit for Elite Spaces",
        GREEN,
    ),
    (
        "Substack (curated list)",
        "Yes — [pub].substack.com/feed per publication",
        "Partial",
        "cityhappenings.substack.com | coolstuffnyc.substack.com | onefinedaynyc.substack.com | blankmanlist.com | preoccupiednyc.substack.com",
        "Monthly NYC guides with curated event picks; weekly social calendars; lifestyle + beauty event mentions",
        "Not a structured calendar — events are embedded in newsletter prose; RSS feeds exist but require NLP/AI parsing to extract event data; best used for discovery/curation rather than direct automation. Top picks: City Happenings, coolstuff.nyc, One Fine Day NYC, The Blankman List, Preoccupied NYC.",
        YELLOW,
    ),
]

for r_idx, row_data in enumerate(rows, start=2):
    *cells, row_color = row_data
    for c_idx, val in enumerate(cells, start=1):
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        cell.font      = Font(name="Arial", size=9,
                              color=ACCENT if c_idx in (2, 4) else "000000")
        cell.fill      = PatternFill("solid", fgColor=row_color)
        cell.alignment = Alignment(vertical="top", wrap_text=True, horizontal="left")
        cell.border    = bdr
    ws.row_dimensions[r_idx].height = 90

# Legend sheet
lg = wb.create_sheet("Legend")
legend_rows = [
    ("Color",  "Meaning",                                                     HDR_BG,   HDR_FG),
    ("Green",  "Fully automatable — RSS and/or clean scrape confirmed",       GREEN,    "000000"),
    ("Yellow", "Workaround needed — partial access, no RSS, or B2B gating",  YELLOW,   "000000"),
    ("Red",    "Manual only — private/locked site or wrong source",           RED_CELL, "000000"),
]
for r, (label, desc, bg, fg) in enumerate(legend_rows, start=1):
    for c, val in enumerate([label, desc], start=1):
        cell = lg.cell(row=r, column=c, value=val)
        cell.font      = Font(name="Arial", size=9, bold=(r == 1), color=fg)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border    = bdr
    lg.row_dimensions[r].height = 22

lg.column_dimensions["A"].width = 14
lg.column_dimensions["B"].width = 70

ws.freeze_panes = "A2"

out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Elite Spaces — Source Audit.xlsx")
wb.save(out)
print("Saved:", out)
